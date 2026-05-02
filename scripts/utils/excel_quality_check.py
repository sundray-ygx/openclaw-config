#!/usr/bin/env python3
"""
Excel文件质量预检脚本

功能：
1. 检测Excel文件格式和完整性
2. 评估文件质量（分级：优秀/良好/一般/差）
3. 提供处理建议
4. 支持批量检查
"""

import os
import sys
from datetime import datetime
from enum import Enum

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少 openpyxl 库，请安装: pip install openpyxl")
    sys.exit(1)


class QualityLevel(Enum):
    EXCELLENT = "优秀"  # 完整、格式正确、有结构
    GOOD = "良好"       # 基本完整、轻微格式问题
    FAIR = "一般"       # 部分数据缺失、格式有问题
    POOR = "差"         # 严重问题、无法使用


def log(message, level="INFO"):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] [{level}] {message}")


def check_excel_file(file_path):
    """
    检查Excel文件质量

    Returns:
        {
            "file": str,
            "quality": QualityLevel,
            "score": int (0-100),
            "issues": list,
            "suggestions": list,
            "details": dict
        }
    """
    log(f"检查文件: {file_path}")

    result = {
        "file": file_path,
        "quality": QualityLevel.POOR,
        "score": 0,
        "issues": [],
        "suggestions": [],
        "details": {}
    }

    # 基础检查
    if not os.path.exists(file_path):
        result["issues"].append("文件不存在")
        return result

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        result["issues"].append(f"无法打开文件: {str(e)}")
        return result

    # 检查工作表
    if len(wb.sheetnames) == 0:
        result["issues"].append("没有工作表")
        return result

    result["details"]["sheet_count"] = len(wb.sheetnames)
    result["details"]["sheet_names"] = wb.sheetnames

    # 检查第一个工作表（假设是主数据表）
    ws = wb.active
    result["details"]["active_sheet"] = ws.title

    # 计算使用的行数和列数
    max_row = ws.max_row
    max_col = ws.max_column
    result["details"]["dimensions"] = f"{max_row}行 x {max_col}列"

    # 检查数据量
    if max_row < 2:
        result["issues"].append("数据行数过少（可能是空表或只有表头）")
        result["score"] -= 30
    elif max_row < 10:
        result["suggestions"].append("数据行数较少，确认是否为完整数据")
        result["score"] -= 10

    # 检查表头
    if max_row >= 1:
        headers = []
        empty_headers = 0
        for col in range(1, min(max_col + 1, 27)):  # 最多检查A-Z列
            cell_value = ws.cell(1, col).value
            if cell_value is None or str(cell_value).strip() == "":
                empty_headers += 1
            else:
                headers.append(str(cell_value).strip())

        result["details"]["headers"] = headers
        result["details"]["empty_headers_in_row1"] = empty_headers

        # 检查表头质量
        if empty_headers > len(headers) * 0.5:
            result["issues"].append("第一行表头缺失过多，可能不是表头行")
            result["score"] -= 20
        elif empty_headers > 0:
            result["suggestions"].append(f"第一行有 {empty_headers} 个空单元格，确认是否为表头")

        # 检查重复表头
        if len(headers) != len(set(headers)):
            duplicates = [h for h in headers if headers.count(h) > 1]
            result["issues"].append(f"存在重复表头: {set(duplicates)}")
            result["score"] -= 15
            result["suggestions"].append("建议合并重复列或重命名")

    # 检查数据完整性
    if max_row >= 2:
        empty_rows = 0
        empty_cells = 0
        total_cells = 0

        for row in range(2, min(max_row + 1, 101)):  # 最多检查前100行数据
            row_empty = True
            for col in range(1, min(max_col + 1, 27)):
                total_cells += 1
                cell_value = ws.cell(row, col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    row_empty = False
                else:
                    empty_cells += 1
            if row_empty:
                empty_rows += 1

        result["details"]["empty_rows_in_data"] = empty_rows
        result["details"]["empty_cell_ratio"] = empty_cells / total_cells if total_cells > 0 else 0

        if empty_rows > max_row * 0.3:
            result["issues"].append(f"数据中空行过多: {empty_rows}/{max_row-1}")
            result["score"] -= 15
            result["suggestions"].append("建议删除空行或检查数据是否完整")

        if empty_cells / total_cells > 0.5:
            result["issues"].append(f"空单元格比例过高: {empty_cells/total_cells:.1%}")
            result["score"] -= 20

    # 检查公式
    formula_count = 0
    if max_row <= 1000:  # 大文件跳过公式检查
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 100), max_col=min(max_col, 26)):
            for cell in row:
                if cell.data_type == 'f':  # formula
                    formula_count += 1

    result["details"]["formula_count"] = formula_count
    if formula_count > 0:
        result["suggestions"].append(f"文件包含 {formula_count} 个公式，确认是否需要保留计算值")

    # 检查合并单元格
    merged_ranges = len(ws.merged_cells.ranges)
    result["details"]["merged_cells"] = merged_ranges
    if merged_ranges > 0:
        result["suggestions"].append(f"文件包含 {merged_ranges} 个合并单元格，处理时需要注意")

    # 计算总分
    result["score"] = max(0, min(100, result["score"] + 50))  # 基础分50

    # 确定质量等级
    if result["score"] >= 85:
        result["quality"] = QualityLevel.EXCELLENT
    elif result["score"] >= 70:
        result["quality"] = QualityLevel.GOOD
    elif result["score"] >= 50:
        result["quality"] = QualityLevel.FAIR
    else:
        result["quality"] = QualityLevel.POOR

    log(f"质量等级: {result['quality'].value} (得分: {result['score']})")

    wb.close()
    return result


def batch_check(directory, pattern="*.xlsx"):
    """批量检查目录下的Excel文件"""
    import glob

    files = glob.glob(os.path.join(directory, pattern))
    if not files:
        log(f"目录 {directory} 中没有找到Excel文件")
        return []

    results = []
    for file_path in files:
        result = check_excel_file(file_path)
        results.append(result)

    return results


def print_report(results):
    """打印检查报告"""
    if not results:
        return

    log("=" * 60)
    log("Excel文件质量检查报告")
    log("=" * 60)

    # 统计
    quality_counts = {level: 0 for level in QualityLevel}
    for result in results:
        quality_counts[result["quality"]] += 1

    log("\n质量分布:")
    for level in QualityLevel:
        count = quality_counts[level]
        if count > 0:
            log(f"  {level.value}: {count} 个")

    log("\n详细结果:")
    log("-" * 60)

    for result in results:
        status_icon = {
            QualityLevel.EXCELLENT: "✓",
            QualityLevel.GOOD: "○",
            QualityLevel.FAIR: "△",
            QualityLevel.POOR: "✗"
        }[result["quality"]]

        log(f"\n{status_icon} {os.path.basename(result['file'])}")
        log(f"  质量等级: {result['quality'].value} (得分: {result['score']})")
        log(f"  规模: {result['details'].get('dimensions', 'N/A')}, 工作表: {result['details'].get('sheet_count', 0)}个")

        if result["issues"]:
            log("  问题:")
            for issue in result["issues"]:
                log(f"    ✗ {issue}")

        if result["suggestions"]:
            log("  建议:")
            for suggestion in result["suggestions"]:
                log(f"    → {suggestion}")

    log("\n" + "=" * 60)


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("用法: python excel_quality_check.py <文件路径/目录> [--batch]")
        print("\n示例:")
        print("  python excel_quality_check.py data.xlsx")
        print("  python excel_quality_check.py ./excels/ --batch")
        sys.exit(1)

    target = sys.argv[1]
    batch_mode = "--batch" in sys.argv

    log("=" * 50)
    log("Excel文件质量预检")
    log("=" * 50)

    if batch_mode and os.path.isdir(target):
        results = batch_check(target)
    elif os.path.isfile(target):
        result = check_excel_file(target)
        results = [result]
    else:
        log(f"目标不存在: {target}", "ERROR")
        return 1

    print_report(results)

    # 返回状态码（如果有"差"质量的文件，返回1）
    has_poor = any(r["quality"] == QualityLevel.POOR for r in results)
    return 1 if has_poor else 0


if __name__ == "__main__":
    sys.exit(main())
